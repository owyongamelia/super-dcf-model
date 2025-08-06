"""
Replacement FastAPI application for merging DCF models with consensus and
public company worksheets.  This module is designed to run on a
platform such as Render and implements the `/upload` endpoint expected
by the provided frontend.  See the module level docstring in
`main.py` for a detailed description of the behaviour.

The primary difference compared to the original `main.py` is that
`generate_output_file` now actually combines the uploaded Excel
workbooks.  It retains formulas and formatting by copying cells and
styles with :mod:`openpyxl` rather than simply returning the
pre‑packaged template.  The consensus workbook is always required and
the optional profile workbook adds an additional ``Public Company``
sheet if present.
"""

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from copy import copy
from openpyxl.cell.cell import MergedCell
from tempfile import NamedTemporaryFile
import shutil
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/upload")
async def upload(consensus: UploadFile = File(...), profile: UploadFile = File(None)):
    """Handle file uploads from the frontend and return the merged workbook.

    Parameters
    ----------
    consensus : UploadFile (required)
        The user provided Excel workbook containing at least a ``Consensus``
        worksheet and optionally a ``Public Company`` worksheet.
    profile : UploadFile (optional)
        An additional workbook used to source the ``Public Company`` sheet
        if it is not present in the consensus workbook.

    Returns
    -------
    StreamingResponse
        A streaming response containing the merged XLSX file with
        appropriate headers for download.
    """
    # Save the uploaded files temporarily on disk.  Doing so allows
    # openpyxl to read them multiple times and simplifies clean‑up.
    consensus_path = "temp_consensus.xlsx"
    with open(consensus_path, "wb") as f:
        shutil.copyfileobj(consensus.file, f)

    profile_path = None
    if profile is not None:
        profile_path = "temp_profile.xlsx"
        with open(profile_path, "wb") as f:
            shutil.copyfileobj(profile.file, f)

    try:
        # Generate the combined workbook
        output_path = generate_output_file(consensus_path, profile_path)

        # Return as a streaming response.  Set the content disposition
        # filename to something indicative for the user.
        return StreamingResponse(
            open(output_path, "rb"),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Merged_DCF_Model.xlsx"},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        # Remove temporary input files
        if os.path.exists(consensus_path):
            os.remove(consensus_path)
        if profile_path and os.path.exists(profile_path):
            os.remove(profile_path)


def generate_output_file(consensus_path: str, profile_path: str | None) -> str:
    """Create a merged Excel workbook containing DCF Model, Consensus and Public Company.

    This function performs the core logic of combining the provided
    consensus workbook with the pre‑packaged DCF model and an optional
    profile workbook.  It preserves formulas, formatting, merged
    ranges, row/column sizes and other basic worksheet settings.  The
    resulting workbook has the ``DCF Model`` sheet first followed by
    ``Consensus`` and, if provided, ``Public Company``.

    Parameters
    ----------
    consensus_path : str
        Path to the uploaded consensus workbook on disk.
    profile_path : str or None
        Path to the uploaded profile workbook on disk, or None if
        omitted.

    Returns
    -------
    str
        Filesystem path to the newly created XLSX file.  The caller
        should delete this file after sending it to the user.
    """
    # Load the template workbook containing the DCF model.  Use
    # data_only=False so that formulas are not converted to values.
    template_wb = load_workbook("DCF Model.xlsx", data_only=False)
    # We'll build our output workbook starting from the template.
    output_wb = template_wb

    # Helper function to copy a worksheet into the output workbook.  We
    # copy cell values/formulas and styles, merged cells, freeze panes,
    # and row/column dimensions.  More advanced features (charts,
    # images, macros) are not handled.
    def copy_sheet(source_sheet, target_title):
        # Remove any existing sheet with the same name in the output to
        # ensure formulas reference the correct sheet.
        if target_title in output_wb.sheetnames:
            std = output_wb[target_title]
            output_wb.remove(std)
        # Create the new sheet
        target_sheet = output_wb.create_sheet(title=target_title)
        # Copy column dimensions
        for col_letter, col_dim in source_sheet.column_dimensions.items():
            new_dim = target_sheet.column_dimensions[col_letter]
            new_dim.width = col_dim.width
            new_dim.hidden = col_dim.hidden
            new_dim.outlineLevel = col_dim.outlineLevel
            # Style may be None if not set; copy only when present
            if col_dim.style:
                new_dim.style = copy(col_dim.style)
        # Copy row dimensions
        for row_idx, row_dim in source_sheet.row_dimensions.items():
            new_dim = target_sheet.row_dimensions[row_idx]
            new_dim.height = row_dim.height
            new_dim.hidden = row_dim.hidden
            new_dim.outlineLevel = row_dim.outlineLevel
            if row_dim.style:
                new_dim.style = copy(row_dim.style)
        # Copy freeze panes
        target_sheet.freeze_panes = source_sheet.freeze_panes
        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        # Copy the cells
        for row in source_sheet.iter_rows():
            for cell in row:
                # Skip phantom cells created as part of a merged range.  Only
                # the top‑left cell of a merged range contains the actual
                # value; other cells of the range are of type MergedCell and
                # assigning to them raises an AttributeError.  We rely on
                # copying the merged ranges below to recreate the merge in
                # the target sheet.
                if isinstance(cell, MergedCell):
                    continue
                new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
                if cell.hyperlink:
                    new_cell.hyperlink = copy(cell.hyperlink)
                if cell.comment:
                    new_cell.comment = copy(cell.comment)
        # Copy printing options.  The sheet_view attribute on a worksheet
        # has no public setter in openpyxl so it cannot be assigned.
        target_sheet.page_setup = copy(source_sheet.page_setup)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        target_sheet.print_options = copy(source_sheet.print_options)

    # Load the consensus workbook
    consensus_wb = load_workbook(consensus_path, data_only=False)
    # Determine which sheets to copy from consensus
    to_copy = []
    if "Consensus" in consensus_wb.sheetnames:
        to_copy.append("Consensus")
    # If the user didn't provide a profile, copy Public Company from consensus if present
    if profile_path is None and "Public Company" in consensus_wb.sheetnames:
        to_copy.append("Public Company")
    # Copy the selected sheets
    for sheet_name in to_copy:
        copy_sheet(consensus_wb[sheet_name], sheet_name)
    # If profile file is provided, copy its Public Company sheet (if exists)
    if profile_path:
        profile_wb = load_workbook(profile_path, data_only=False)
        if "Public Company" in profile_wb.sheetnames:
            copy_sheet(profile_wb["Public Company"], "Public Company")
    # Write the output workbook to a temporary file
    temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    output_wb.save(temp_file.name)
    return temp_file.name
